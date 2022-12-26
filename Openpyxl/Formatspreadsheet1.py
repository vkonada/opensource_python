# import openpyxl module
import openpyxl

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute.
sheet = wb.active

# writing to the specified cell
sheet.cell(row = 1, column = 1).value = ' hello '

sheet.cell(row = 2, column = 2).value = ' everyone '

# set the height of the row
sheet.row_dimensions[1].height = 70

# set the width of the column
sheet.column_dimensions['B'].width = 20

# save the file
wb.save('sample.xlsx')

#################


import openpyxl

# import Font function from openpyxl
from openpyxl.styles import Font


wb = openpyxl.Workbook()
sheet = wb.active

sheet.cell(row = 1, column = 1).value = "GeeksforGeeks"

# set the size of the cell to 24
sheet.cell(row = 1, column = 1).font = Font(size = 24 )

sheet.cell(row = 2, column = 2).value = "GeeksforGeeks"

# set the font style to italic
sheet.cell(row = 2, column = 2).font = Font(size = 24, italic = True)

sheet.cell(row = 3, column = 3).value = "GeeksforGeeks"

# set the font style to bold
sheet.cell(row = 3, column = 3).font = Font(size = 24, bold = True)

sheet.cell(row = 4, column = 4).value = "GeeksforGeeks"

# set the font name to 'Times New Roman'
sheet.cell(row = 4, column = 4).font = Font(size = 24, name = 'Times New Roman')

wb.save('sample.xlsx')

