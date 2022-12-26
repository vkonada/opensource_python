import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active

# merge cell from A2 to D4 i.e.
# A2, B2, C2, D2, A3, B3, C3, D3, A4, B4, C4 and D4 .
# A2:D4' merges 12 cells into a single cell.
sheet.merge_cells('A2:D4')

sheet.cell(row = 2, column = 1).value = 'Twelve cells join together.'

# merge cell C6 and D6
sheet.merge_cells('C6:D6')

sheet.cell(row = 6, column = 6).value = 'Two merge cells.'

wb.save('sample.xlsx')

############

import openpyxl


wb = openpyxl.load_workbook('sample.xlsx')
sheet = wb.active

# unmerge the cells
sheet.unmerge_cells('A2:D4')

sheet.unmerge_cells('C6:D6')

wb.save('sample.xlsx')
