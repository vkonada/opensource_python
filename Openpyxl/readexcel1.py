# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "gfg.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row = 1, column = 1)

# Print value of cell object
# using the value attribute
print(cell_obj.value)

####################
####################
# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "gfg.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

# printing the value of first column
# Loop will print all values
# of first column
print("\nValue of first column")
for i in range(1, row + 1):
	cell_obj = sheet_obj.cell(row = i, column = 1)
	print(cell_obj.value)
	
# printing the value of first column
# Loop will print all values
# of first row
print("\nValue of first row")
for i in range(1, column + 1):
	cell_obj = sheet_obj.cell(row = 2, column = i)
	print(cell_obj.value, end = " ")

  
###########
###########
  
# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "gfg.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj['A1': 'B6']

# Print value of cell object
# using the value attribute
for cell1, cell2 in cell_obj:
	print(cell1.value, cell2.value)

  
